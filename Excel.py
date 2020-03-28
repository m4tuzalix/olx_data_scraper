from openpyxl import Workbook, load_workbook
from datetime import datetime
class Excel():
    def __init__(self):
        self.date = datetime.now().date()
        self.wb = Workbook()
        self.ws = self.wb.active
        self.workbook_name = f"({self.date}).xlsx"
        self.just_created = False #// to avoid double check after file creation
        try:
            self.load = load_workbook(self.workbook_name)
        except FileNotFoundError:
            headers = ["link", "nazwa", "lokacja", "powierzchnia", "czynsz", "cena", "opis", "zdjecia"]
            for x in range(len(headers)):
                self.ws.cell(row=1, column=x+1, value=headers[x]) #// +1 because it must be at least 1
            self.wb.save(self.workbook_name)
            self.just_created = True
      
    def add_data(self, data):
        self.load = load_workbook(self.workbook_name)
        self.ns = self.load.get_sheet_by_name("Sheet")
        if self.just_created != True:
            sheet = self.load.get_sheet_by_name("Sheet")
            max_row = sheet.max_row
            link = data[0]
            found = False
            for x in range(1, max_row+1):
                    cell = sheet.cell(row=x+1, column=1) #// column or row must be at least 1
                    if str(cell.value) == str(link):
                        found = True
                        break
            if found == False:
                self.ns.append(data)
        else:
            self.ns.append(data)
        self.load.save(self.workbook_name)


